import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import pandas as pd
from PIL import Image
from unidecode import unidecode
import math
import json
import numpy as np
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText
from io import BytesIO
import zipfile

# Cargar la imagen
logo_image = Image.open("LOGO_BUKZ.webp")  # Reemplaza con la ruta o nombre de archivo correcto

resized_image = logo_image.resize((200, 50))  # Especifica las dimensiones deseadas

# Mostrar la imagen redimensionada en la barra lateral
st.sidebar.image(resized_image)


with st.sidebar:
    choose = option_menu("Menú de opciones", ['Actualización de inventario celesa', 'Creación de productos', 'Corte Proveedores'],
    icons=["list check", "database up", "check2 square"], menu_icon="cast", default_index=0,
    styles={ "container": {"padding": "5!important", "background-color": "#fafafa"},
        "icon": { "font-size": "25px"}, 
        "nav-link": {"font-size": "16px", "text-align": "left", "margin":"0px", "--hover-color": "#eee", "color": "black"},
        "nav-link-selected": {"background-color": "#F9CB00"}})

if 'smtp_user' not in st.session_state:
    st.session_state['smtp_user'] = ''
if 'smtp_password' not in st.session_state:
    st.session_state['smtp_password'] = ''
if 'archivo_cargado' not in st.session_state:
    st.session_state['archivo_cargado'] = False
if 'procesado' not in st.session_state:
    st.session_state['procesado'] = False


if choose == 'Actualización de inventario celesa':
    st.title("Actualización de inventario celesa")
    st.write("Cargar archivos CSV:")
    #st.set_option('deprecation.showfileUploaderEncoding', False)  # Evita el aviso de codificación

    st.markdown("<h3>Archivo Productos</h3>", unsafe_allow_html=True)
    uploaded_file1 = st.file_uploader("El archivo de producto debe tener las columnas: ID,  Variant ID,  Vendor,  Variant SKU,  Variant Barcode,  Inventory Available Dropshipping [España]", type=["csv"], key="archivo_productos")

    st.markdown("<h3>Archivo Azeta</h3>", unsafe_allow_html=True)
    uploaded_file2 = st.file_uploader("", type=["csv"], key="archivo_azeta")


    if uploaded_file1 is not None and uploaded_file2 is not None:
        st.write("Presiona el botón para continuar")
        if st.button("Continuar"):
            info_placeholder = st.empty()
            info_placeholder.info("Cargando...")

            df_products = pd.read_csv(uploaded_file1)
            df_azeta = pd.read_csv(uploaded_file2, sep=';', header=None)
            df_azeta.columns = ['Variant SKU', 'Stock_Azeta']

            try:
                df_products = df_products.loc[df_products['Vendor'] == 'Bukz España']
                df_products = df_products[['ID', 'Variant ID', 'Vendor', 'Variant SKU', 'Variant Barcode', 'Inventory Available: Dropshipping [España]']]
                df_products.insert(1, 'Command', 'UPDATE')
                
                df_merged = pd.merge(df_products, df_azeta, on="Variant SKU", how='left')
                df_merged['Inventory Available: Dropshipping [España]'].fillna(0, inplace=True)
                df_merged['Stock_Azeta'].fillna(0, inplace=True)
                df_merged['Stock_Azeta'] = df_merged['Stock_Azeta'].astype(int)
                df_merged['Inventory Available: Dropshipping [España]'] = pd.to_numeric(df_merged['Inventory Available: Dropshipping [España]'], errors='coerce').fillna(0).astype(int)

                df_merged['Inventory Available: Dropshipping [España]'] = df_merged['Inventory Available: Dropshipping [España]'].astype(int)

                comparar_filas = lambda x: 1 if x['Inventory Available: Dropshipping [España]'] == x['Stock_Azeta'] else 0
                df_merged['Resultado'] = df_merged.apply(comparar_filas, axis=1)
                df_final = df_merged.loc[df_merged['Resultado'] == 0]
                df_final['Inventory Available: Dropshipping [España]'] = df_final['Stock_Azeta']
                            
                
                df_final.drop(['Stock_Azeta', 'Resultado'], axis=1, inplace=True)
                df_final = df_final.astype({'ID':str, 'Variant ID':str, 'Vendor':str, 'Variant SKU':str, 
                                           'Variant Barcode':str, 'Inventory Available: Dropshipping [España]':str})
                
                info_placeholder.empty()
                st.write(df_final)

                # Botón de descarga sin base64
                st.download_button(
                    label="Descargar CSV",
                    data=df_final.to_csv(index=False),
                    file_name="resultado_cruzado.csv",
                    mime="text/csv"
                )
            except Exception as e:
                info_placeholder.empty()
                st.error(f"Error: {str(e)}")
                st.error(f"Traceback: {traceback.format_exc()}")
    else:
        st.info("Por favor, carga ambos archivos para continuar.")


elif choose == 'Creación de productos':
    st.title("Creación de productos")
    st.markdown("<h3>Plantilla creación de productos</h3>", unsafe_allow_html=True)
   
    #st.set_option('deprecation.showfileUploaderEncoding', False)  # Evita el aviso de codificación

# Ruta del archivo local
    file_path = "plantilla_creacion_productos.xlsx"
    
    # Botón de descarga
    def download_file():
        with open(file_path, "rb") as file:
            btn = st.download_button(label="Descargar archivo", data=file, file_name="plantilla_creacion_productos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return btn
    # Llamar a la función para mostrar el botón de descarga
    download_file()
    
    st.write("Cargar archivo XLSX (Formato Excel):")
    st.markdown("<h3>Archivo con datos</h3>", unsafe_allow_html=True)
    uploaded_file3 = st.file_uploader("Los nombres de las columnas del archivo debe conservar el mismo nombre de la plantilla de creación de productos", type=["xlsx"], key="archivo_productos")
    if uploaded_file3 is not None:
        st.write("Presiona el botón para continuar")
        if st.button("Continuar"):
            info_placeholder = st.empty()
            info_placeholder.info("Cargando...")

            df2 = pd.read_excel(uploaded_file3, sheet_name='Products', engine='openpyxl')

            try:  
                df = pd.DataFrame()
                df['Title'] = df2['Titulo'].str.title()
                
                df['Command'] = 'NEW'
                
                df['Body HTML'] = df2['Sipnosis']
                
                df2['Variant SKU'] = df2['SKU'].apply(lambda x: str(x).replace('.0', ''))
                
                df['Handle'] = df['Title'].str.lower().apply(unidecode).str.replace(r'[^\w\s]+', '', regex=True).str.replace(' ', '-') + '-' + df2['Variant SKU']
                
                #Vendor
                df['Vendor'] = df2['Vendor']
                
                #Vendor
                df['Type'] = 'Libro'
                
                #tags
                df['Tags'] = pd.Series(dtype=str)
                
                #Status
                df['Status'] = 'Active'
                
                #Published
                df['Published'] = 'TRUE'
                
                #Published Scope
                df['Published Scope'] = 'global'
                
                #Gift Card
                df['Gift Card'] = 'FALSE'
                
                #Row #
                df['Row #'] = 1
                
                #Top Row
                df['Top Row'] = 'TRUE'
                
                # Option1 Name
                df['Option1 Name'] = 'Title'
                
                #Option1 Value
                df['Option1 Value'] = 'Default Title'
                
                # Option2 Name
                df['Option2 Name'] = pd.Series(dtype=str)
                
                # Option2 Value
                df['Option2 Value'] = pd.Series(dtype=str)
                
                # Option3 Name
                df['Option3 Name'] = pd.Series(dtype=str)
                
                # Option3 Value
                df['Option3 Value'] = pd.Series(dtype=str)
                
                #Variant Position
                df['Variant Position'] = pd.Series(dtype=str)
                
                #Variant SKU
                df['Variant SKU'] = df2['SKU'].apply(lambda x: str(x).replace('.0', ''))
                #Variant Barcode
                df['Variant Barcode'] = df['Variant SKU']
                
                #Variant Image
                df['Image Src'] = df2['Portada (URL)']
                
                df['Variant Price'] = df2["Precio"]
                
                df['Variant Compare At Price'] = df2["Precio de comparacion"]
                
                #Variant Taxable
                df['Variant Taxable'] = 'FALSE'
                
                #Variant Tax Code
                df['Variant Tax Code'] = pd.Series(dtype=str)
                
                #Variant Inventory Tracker
                df['Variant Inventory Tracker'] = 'shopify'
                
                #Variant Inventory Policy
                df['Variant Inventory Policy'] = 'deny'
                
                #Variant Inventory Tracker
                df['Variant Fulfillment Service'] = 'manual'
                
                #Variant Requires Shipping
                df['Variant Requires Shipping'] = 'TRUE'
                #Variant Weight
                df['Variant Weight'] = df2['peso (kg)']
                
                #Variant Weight Unit
                df['Variant Weight Unit'] = df['Variant Weight'].apply(lambda x: 'kg' if pd.notnull(x) else np.nan)
                
                #Metafield: custom.autor [single_line_text_field]
                df['Metafield: custom.autor [single_line_text_field]'] = df2["Autor"].fillna("").str.title()
                
                replacements2 = {'Español' : '["Español"]', 'Ingles' : '["Ingles"]', 'Frances' : '["Frances"]', 'Italiano' : '["Italiano"]', 'Portugues' : '["Portugues"]', 
                    'Aleman' : '["Aleman"]', 'Bilingue (Español-Ingles)' : '["Bilingue (Español-Ingles)"]', 'Bilingue (Español-Portugues)' : '["Bilingue (Español-Portugues)"]', 
                    'Vasco' : '["Vasco"]', 'Gallego' : '["Gallego"]', 'Latin' : '["Latin"]', 'Ruso' : '["Ruso"]', 'Arabe' : '["Arabe"]', 'Chino' : '["Chino"]', 
                    'Japones' : '["Japones"]', 'Catalan' : '["Catalan"]', 'Rumano' : '["Rumano"]', 'Holandes' : '["Holandes"]', 'Bulgaro' : '["Bulgaro"]', 'Griego' : '["Griego"]', 
                    'Polaco' : '["Polaco"]', 'Checo' : '["Checo"]', 'Sueco' : '["Sueco"]'}
                
                df['Metafield: custom.idioma [list.single_line_text_field]'] = df2['Idioma'].apply(lambda x: replacements2.get(x, x))
                
                replacements = {
                    'Tapa Dura' : '["Tapa Dura"]', 'Tapa Blanda' : '["Tapa Blanda"]', 'Bolsillo' : '["Bolsillo"]', 'Libro de lujo' : '["Libro de lujo"]', 
                    'Espiral' : '["Espiral"]', 'Tela' : '["Tela"]', 'Grapado' : '["Grapado"]', 'Fasciculo Encuadernable' : '["Fasciculo Encuadernable"]', 
                    'Troquelado' : '["Troquelado"]', 'Anillas' : '["Anillas"]', 'Otros' : '["Otros"]'}
                
                df['Metafield: custom.formato [list.single_line_text_field]'] = df2['Formato'].apply(lambda x: replacements.get(x, x))
                
                df['Metafield: custom.alto [dimension]'] = df2['Alto'].apply(lambda x: np.nan if np.isnan(x) else json.dumps({"value": x, "unit": "cm"}))
                
                df['Metafield: custom.ancho [dimension]'] = df2['Ancho'].apply(lambda x: np.nan if np.isnan(x) else json.dumps({"value": x, "unit": "cm"}))
                
                df['Metafield: custom.editorial [single_line_text_field]'] = df2['Editorial'].str.title()
                
                df['Metafield: custom.numero_de_paginas [number_integer]'] = df2['Numero de paginas']
                
                df['Metafield: custom.ilustrador [single_line_text_field]'] = df2['Ilustrador']
                
                df['Image Alt Text'] = 'Libro ' + df['Title']+' '+ df['Variant SKU']
                df_crear = df
                
                info_placeholder.empty()
                st.write(df_crear)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_crear.to_excel(writer, index=False)
        
                # Mostrar el botón de descarga
                st.download_button(
                    label="Descargar archivo xlsx",
                    data=output.getvalue(),
                    file_name="resultado_crear_productos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                
            except Exception as e:
                info_placeholder.empty()
                st.error(f"Error: {str(e)}")
    else:
        st.info("Por favor, carga el archivo para continuar.")
    

elif choose == 'Corte Provedores':
    
    st.title("Corte Provedores")
    proveedores_df = pd.read_excel("Proveedores Corte.xlsx")
    
    if 'envio_proveedores' not in st.session_state or st.session_state['envio_proveedores'].empty:
        st.session_state['envio_proveedores'] = pd.DataFrame(columns=['Proveedor', 'Estado'])

    st.markdown("<h3>Archivo de ventas mensuales</h3>", unsafe_allow_html=True)
    uploaded_file_ventas_mesuales = st.file_uploader("", type=["xlsx"], key="archivo_productos")
    
    if uploaded_file_ventas_mesuales is not None:
        if st.button("Continuar"):
            st.session_state['archivo_cargado'] = True

    if st.session_state['archivo_cargado']:
        
        grouped_dfs = None  # Inicializa la variable aquí

        # Widgets para ingresar datos del usuario
        st.session_state.smtp_user = st.text_input("Ingrese el usuario SMTP:", value=st.session_state.smtp_user)
        st.session_state.smtp_password = st.text_input("Ingrese la contraseña SMTP:", type="password", value=st.session_state.smtp_password)
 
        mes = st.text_input("Ingrese el mes:")  # No hay valor predeterminado
        año = st.text_input("Ingrese el año:")  # No hay valor predeterminado
        
        if 'nombre_remitente' not in st.session_state:
            st.session_state['nombre_remitente'] = "Sebastian Barrios - Bukz"  # Valor por defecto
 
        #Otros campos de texto
        remitente_default = "Sebastian Barrios - Bukz"
        remitente_otros = "Otro (escriba abajo)"
        nombre_remitente_seleccion = st.selectbox("Seleccione el nombre del remitente:", [remitente_default, remitente_otros])

        # Actualizar el valor de nombre_remitente en st.session_state
        if nombre_remitente_seleccion == remitente_otros:
            if 'nombre_remitente_personalizado' not in st.session_state:
                st.session_state['nombre_remitente_personalizado'] = ''
            st.session_state['nombre_remitente'] = st.text_input("Ingrese el nombre del remitente:", key="nombre_remitente_personalizado")
        else:
            st.session_state['nombre_remitente'] = remitente_default
            
        # Firma
        firma_default = "Sebastian Barrios - Analista de Operaciones"
        firma_otros = "Otra (escriba abajo)"
        firma_seleccion = st.selectbox("Seleccione la firma:", [firma_default, firma_otros])
        
        if st.button("Procesar archivo y enviar correos"):
            
            df = pd.read_excel(uploaded_file_ventas_mesuales)
            df['pos_location_name'] = df['pos_location_name'].replace({
                'Bukz Las Lomas': 'Medellin',
                'Bukz Tesoro': 'Medellin',
                'Bukz Mattelsa': 'Medellin',
                '': 'Medellin',
                'Bukz St. Patrick': 'Bogota'})
            
            df['pos_location_name'] = df['pos_location_name'].fillna('Medellin')
            proveedores_a_eliminar = ["603 La Gran Via", "Alejandra Márquez Villegas", "Alejandro Salazar Yusti",
                "Andina", "Books for U", "Bukz", "Bukz España", "Bukz USA", "Bukz.co",
                "Fernando Ayerbe", "Grupo Editorial Planeta", "Juan D. Hoyos Distribuciones SAS",
                "Libros de Ruta", "Luminosa", "Melon", "Penguin RandomHouse",
                "Pergamino Café", "Postobon", "Tea market", "Torrealta", "Urban",
                "Álvaro González Alorda"]

            df = df[~df['product_vendor'].isin(proveedores_a_eliminar)]
            
            # Función personalizada para sumar teniendo en cuenta la lógica de negativos y positivos
            def custom_sum(group):
                if all(group <= 0):  # Si todos los valores son negativos o cero, suma todos
                    return group.sum()
                else:  # Si no, suma solo los positivos y ceros
                    return group[group >= 0].sum()
                
            # Aplicar la función personalizada para la agrupación
            df_grouped = df.groupby(['product_title', 'product_vendor', 'variant_sku', 'pos_location_name'])['net_quantity'].apply(custom_sum).reset_index()
            
            def format_variant_sku(value):
                try:
                    return '{:.0f}'.format(float(value))
                except ValueError:
                    return value  # Devuelve el valor original si no se puede convertir
            
            df_grouped['variant_sku'] = df_grouped['variant_sku'].apply(format_variant_sku) 
            grouped_dfs = {vendor: sub_df for vendor, sub_df in df_grouped.groupby('product_vendor')}
            
            # Configuración del servidor de correo
            smtp_server = 'smtp.gmail.com'
            smtp_port = 587                
            # Conexión al servidor SMTP
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(st.session_state.smtp_user, st.session_state.smtp_password)

            imagen_url = "https://ci3.googleusercontent.com/mail-sig/AIorK4zk7DTZK_4Nl0qLnpmzJnoAhaN3t08JpWQmDUdtbhe-nJySTGmVsdjlqZr7sVzEJzCFTSGzHY8" 
            
            # Función para ajustar el ancho de las columnas
            def ajustar_ancho_columnas(workbook, ancho):
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    for col in worksheet.columns:
                        col_letter = get_column_letter(col[0].column)
                        worksheet.column_dimensions[col_letter].width = ancho
            
            # Función modificada para enviar un DataFrame como archivo Excel con múltiples hojas a múltiples correos
            def enviar_df_por_correo(df, emails, subject, body_message,nombre_remitente):
                # Crear un buffer en memoria para el archivo Excel
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for ciudad in ['Medellin', 'Bogota']:
                        df_ciudad = df[df['pos_location_name'] == ciudad]
                        if not df_ciudad.empty:
                            df_ciudad.to_excel(writer, sheet_name=ciudad, index=False)
                    # No es necesario llamar a save() aquí, se maneja automáticamente.
            
                # Mover el puntero al inicio del buffer
                output.seek(0)
            
                # Crear el mensaje de correo electrónico
                msg = MIMEMultipart()
                msg['From'] = f"{nombre_remitente}"
                msg['To'] = ", ".join(emails)
                msg['Subject'] = subject
            
                # Agregar el cuerpo del mensaje
                msg.attach(MIMEText(body_message, 'html'))
            
                # Adjuntar el archivo Excel desde el buffer en memoria
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(output.getvalue())  # Usamos getvalue() para obtener el contenido del buffer
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment', filename=f'Corte_Ventas_{mes}.xlsx')
                msg.attach(part)
            
                # Enviar el correo
                server.send_message(msg)
            
                # Cerrar el buffer
                output.close()
                
                st.session_state['procesado'] = True
            
            # Texto del cuerpo del correo
            body_message = f"""<p>Buenas tardes,</p>
            
            <p>Espero que estés muy bien. Adjunto información del corte del mes de {mes}.</p>
            
            <p>Le recordamos que la cuenta de cobro o factura debe ser remitida <strong>exclusivamente a facturacion@bukz.co antes del día 25 del mes en curso</strong>.</p>
            
            <p>Para cualquier consulta o asunto adicional, no dude en contactar a los siguientes departamentos:</p>
            <ul>
                <li>Cartera: cartera@bukz.co</li>
                <li>Bodega y Devoluciones: cedi@bukz.co</li>
                <li>Área Comercial: malu@bukz.co</li>
            </ul>
            
            <p>Saludos cordiales,</p>
            
            <p><strong style="color: gray;">{firma_seleccion}</strong></p>
            <p><img src="{imagen_url}" alt="Logo Bukz" style="width: 150px;"></p>  
            
            """
            
            # Enviar correos a proveedores
            for proveedor, sub_df in grouped_dfs.items():
                correos_filtrados = proveedores_df[proveedores_df['Proveedores'] == proveedor]['Correo Medellin']
                if not correos_filtrados.empty:
                    correos = correos_filtrados.iloc[0].split(';')
                    try:
                        enviar_df_por_correo(sub_df, correos, f"Corte {mes} {año} - {proveedor}", body_message, st.session_state['nombre_remitente'])
                        nuevo_registro = {'Proveedor': proveedor, 'Estado': 'Enviado'}
                    except Exception as e:
                        nuevo_registro = {'Proveedor': proveedor, 'Estado': f'Fallo en el envío: {e}'}
                else:
                    nuevo_registro = {'Proveedor': proveedor, 'Estado': 'Correo no encontrado'}
            
                # Aquí usamos append y luego reasignamos el DataFrame actualizado
                nuevo_registro_df = pd.DataFrame([nuevo_registro])
                st.session_state['envio_proveedores'] = pd.concat([st.session_state['envio_proveedores'], nuevo_registro_df], ignore_index=True)
                
                

        if st.session_state['procesado']:
            # Mostrar el DataFrame con los resultados de los envíos
            st.dataframe(st.session_state['envio_proveedores'])
            
            # Función para convertir un DataFrame a Excel y luego codificarlo para la descarga
            def to_excel(df):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                processed_data = output.getvalue()
                return processed_data
            
            # Función para guardar un DataFrame en un buffer como archivo Excel
            def guardar_excel_en_buffer(df, nombre_archivo):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Crear una hoja para cada ciudad en el archivo Excel solo si el DataFrame correspondiente no está vacío
                    for ciudad in ['Medellin', 'Bogota']:
                        df_ciudad = df[df['pos_location_name'] == ciudad]
                        if not df_ciudad.empty:
                            df_ciudad.to_excel(writer, sheet_name=ciudad, index=False)
            
                    # Verificar si se agregaron hojas al archivo
                    if not writer.sheets:
                        # Si no hay hojas, agrega una hoja con un DataFrame vacío para evitar el error
                        pd.DataFrame().to_excel(writer, sheet_name='Vacio', index=False)
            
                output.seek(0)
                return output.getvalue()
            
            # Crear un archivo ZIP en memoria
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, False) as zip_file:
                
                # Agregar el DataFrame de estado de envío como archivo Excel en el ZIP
                estado_envio_excel = to_excel(st.session_state['envio_proveedores'])
                zip_file.writestr('estado_envio.xlsx', estado_envio_excel)
            
                # Agregar los archivos Excel de cada proveedor en el ZIP
                if grouped_dfs is not None:
                    for proveedor, sub_df in grouped_dfs.items():
                        excel_data = guardar_excel_en_buffer(sub_df, f'{proveedor}.xlsx')
                        zip_file.writestr(f'{proveedor}.xlsx', excel_data)
            
            # Preparar el buffer para la descarga
            zip_buffer.seek(0)
            
            # Botón para descargar el archivo ZIP que incluye tanto el estado de envío como los archivos de proveedores
            st.download_button(
                label="Descargar Estado de Envíos y Archivos de Proveedores",
                data=zip_buffer,
    file_name="envios_y_proveedores.zip",
    mime='application/zip')
